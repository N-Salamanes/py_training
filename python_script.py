# # # # # IMPORTS # # # # #
import csv
import openpyxl
import datetime


# # # # # CONSTANTS & SHARED VARIABLES # # # # #
# NOTE: {} is a single placeholder where data can be substituted in its place.
VLOOKUP_TEMPLATE = "=VLOOKUP({},{},{},{})"
WEEKNUM_TEMPLATE = "=WEEKNUM({},{})"
MONTH_TEMPLATE   = "=MONTH({})"
YEAR_TEMPLATE    = "=YEAR({})"

# Map the column to their numeric index in the RAW DATA sheet. E.g. Marketplace (column A) = 1
# NOTE: Index for Excel files start at 1.
RDS_MARKETPLACE_COL  = 1
RDS_STATUS_COL       = 2
RDS_COUNTRY_COL      = 3
RDS_PRODUCT_CODE_COL = 4
RDS_PRODUCT_NAME_COL = 5
RDS_SALES_COL        = 6
RDS_DATE_COL         = 7
RDS_WEEK_NUMBER_COL  = 8
RDS_MONTH_COL        = 9
RDS_YEAR_COL         = 10

# Map the column to their numeric index in Marketplace 1 .csv file. Column A = 0, Column N = 13
# NOTE: Index for .csv files start at 0.
MKTP1_STATUS_COL       = 13
MKTP1_COUNTRY_COL      = 2
MKTP1_PRODUCT_CODE_COL = 3
MKTP1_SALES_COL        = 10
MKTP1_DATE_COL         = 1
 
# Put code your in the function, called 'main()', below
def main():
  # A variable (string) which contains the path to the master workbook
  master_wb_path = './master.xlsx'
  # A variable (object) which points to and loads the master workbook
  master_wb = openpyxl.load_workbook(master_wb_path, data_only=False, keep_links=True)
  # A variable (object) which points to RAW DATA sheet, in the master workbook
  raw_data_sht = master_wb['RAW DATA']

  # A variable (integer) which contains what row we're writing to, in the RAW DATA sheet.
  # We're initially setting it to be the first empty row (last row used + 1) in the RAW DATA sheet.
  row_to_write = raw_data_sht.max_row + 1


  # # # # # # # # # # PROCESS THE FIRST MARKET PLACE # # # # # # # # # #
  # A variable (string) which contains the path to the first marketplace .csv file
  marketplace1_path = 'marketplace_1.csv'
  # A variable (object) which points to and loads the .csv file
  marketplace1 = open(marketplace1_path)
  # A variable (object) which points to the rows in the .csv file
  marketplace1_reader = csv.reader(marketplace1)

  # This for-loop will go through the .csv file line by line. The variable `row` used in the for-loop, contains all the data for a single row at a time.
  # The data type for the variable `row` is a List.
  # E.g. ['Enterprise', '1/11/2013', 'Mexico', '123456-03', 'Low', '$1,660.00', '$5.00', '$125.00', '$207,500.00', '$4,150.00', '$203,350.00', '$199,200.00', '$4,150.00', 'Closed']
  # To access the country column, it would be: row[2]
  for row in marketplace1_reader:
    # Depending on how the .csv file is formatted, data may have starting and/or trailing whitespace. Therefore, we need to strip the whitespace by using .strip() on the data
    # NOTE: NOW USE row_cleaned instead of row, when accessing the data within the rows
    row_cleaned = [value.strip() for value in row]
  
    # # # # # CHECK THE FOLLOWING CONDITION(S) BEFORE PROCEEDING TO COPY OVER DATA # # # # #

    # Only copy data if Marketplace 1 status column is 'Closed'. Therefore, if it's anything other than 'Closed', then execute the code in this if statement.
    if row_cleaned[MKTP1_STATUS_COL] != 'Closed':
      # `continue` means to stop this iteration of the for-loop and start the next iteration immediately. The code below `continue`` will not be executed for this iteration.
      continue

    # # # # # START COPYING OVER DATA # # # # #
    
    # Set MARKETPLACE column the name of the marketplace
    raw_data_sht.cell(row=row_to_write, column=RDS_MARKETPLACE_COL).value = 'Marketplace 1'


    # Set STATUS column. raw data sheet(row=row_to_write,column=2) = marketplace 1 (row=row,column=14)
    raw_data_sht.cell(row=row_to_write, column=RDS_STATUS_COL).value = row_cleaned[MKTP1_STATUS_COL]


    # Set COUNTRY column. raw data sheet(row=row_to_write,column=3) = marketplace 1 (row=row,column=2)
    raw_data_sht.cell(row=row_to_write, column=RDS_COUNTRY_COL).value = row_cleaned[MKTP1_COUNTRY_COL]


    # Set PRODUCT CODE column. raw data sheet(row=row_to_write,column=4) = marketplace 1 (row=row,column=3)
    raw_data_sht.cell(row=row_to_write, column=RDS_PRODUCT_CODE_COL).value = row_cleaned[MKTP1_PRODUCT_CODE_COL]


    # Set PRODUCT NAME column. Writes a vlookup formula to the cell
    # Convert column number to its corresponding letter. E.g. 4 = D 
    rds_product_code_col_letter = openpyxl.utils.cell.get_column_letter(RDS_PRODUCT_CODE_COL)

    # First parameter, combine column letter with row to write. E.g 'D' + '18' = 'D18'
    # NOTE: Notice the type conversion of `row_to_write` to a string from an integer? str(row_to_write) converts the variables type to a string.
    product_name_formula_first_param = rds_product_code_col_letter + str(row_to_write)
    product_name_formula_second_param = 'SKU!A:B'
    product_name_formula_third_param = '2'
    product_name_formula_fourth_param = '0' 

    # Combine the formula components together. VLOOKUP formula for PRODUCT NAME
    # e.g: product_name_formula = '=VLOOKUP({},{},{},{})'
    product_name_formula = VLOOKUP_TEMPLATE.format(product_name_formula_first_param, \
      product_name_formula_second_param, \
      product_name_formula_third_param, \
      product_name_formula_fourth_param)

    # Finally, write to the PRODUCT NAME column using the formula crafted above
    raw_data_sht.cell(row=row_to_write, column=RDS_PRODUCT_NAME_COL).value = product_name_formula


    # Set SALES column. raw data sheet(row=row_to_write,column=6) = marketplace 1 (row=row,column=10)
    # We need to remove '$' and ',' from the text and treat it as a number (float), instead of a string
    sales = row_cleaned[MKTP1_SALES_COL]
    sales_cleaned = sales.replace('$','').replace(',','')
    raw_data_sht.cell(row=row_to_write, column=RDS_SALES_COL).value = float(sales_cleaned)


    # Set DATE column. raw data sheet(row=row_to_write,column=7) = marketplace 1 (row=row,column=1)
    date_read = row_cleaned[MKTP1_DATE_COL]
    # Set the date format, being read from the marketplace csv, according to the specifications in the 2nd parameter below.
    date_datatype_formatted = datetime.datetime.strptime(date_read, '%d/%m/%Y')
    # Set the cell DATE column. Notice we're setting the value as a datetime variable type and not a string. If we passed in a string, excel won't treat the cell like a date data type.
    raw_data_sht.cell(row=row_to_write, column=RDS_DATE_COL).value = date_datatype_formatted
    # Format the cell to just a date and not date & time
    raw_data_sht.cell(row=row_to_write, column=RDS_DATE_COL).number_format = 'dd/mm/yyyy' # This line sets the formatting of the cell


    # Set WEEK NUMBER column.
    # Convert column number to its corresponding letter. E.g. 4 = D 
    rds_date_col_letter = openpyxl.utils.cell.get_column_letter(RDS_DATE_COL)
    # First parameter, combine column letter with row to write. E.g 'D' + '18' = 'D18'
    # NOTE: Notice the type conversion of `row_to_write` to a string from an integer? str(row_to_write) converts the variables type to a string.
    week_number_formula_first_param = rds_date_col_letter + str(row_to_write)
    week_number_formula_second_param = '1'

    #Combine the formula components together. VLOOKUP formula for PRODUCT NAME
    # =WEEKNUM({},{})
    week_number_formula = WEEKNUM_TEMPLATE.format(week_number_formula_first_param, week_number_formula_second_param)

    # Finally, write to the WEEK NUMBER column
    raw_data_sht.cell(row=row_to_write, column=RDS_WEEK_NUMBER_COL).value = week_number_formula


    # Set MONTH
    # Convert column number to its corresponding letter. E.g. 4 = D 
    rds_date_col_letter = openpyxl.utils.cell.get_column_letter(RDS_DATE_COL)
    # First parameter, combine column letter with row to write. E.g 'D' + '18' = 'D18'
    # NOTE: Notice the type conversion of `row_to_write` to a string from an integer? str(row_to_write) converts the variables type to a string.
    month_formula_first_param = rds_date_col_letter + str(row_to_write)

    #Combine the formula components together. MONTH formula for MONTH column
    # =MONTH({})
    month_formula = MONTH_TEMPLATE.format(month_formula_first_param)

    # Finally, write to the MONTH column
    raw_data_sht.cell(row=row_to_write, column=RDS_MONTH_COL).value = month_formula


    # PROCESS YEAR
    # Convert column number to its corresponding letter. E.g. 4 = D 
    rds_date_col_letter = openpyxl.utils.cell.get_column_letter(RDS_DATE_COL)
    # First parameter, combine column letter with row to write. E.g 'D' + '18' = 'D18'
    # NOTE: Notice the type conversion of `row_to_write` to a string from an integer? str(row_to_write) converts the variables type to a string.
    year_formula_first_param = rds_date_col_letter + str(row_to_write)

    #Combine the formula components together. YEAR formula for YEAR column
    # =YEAR({})
    year_formula = YEAR_TEMPLATE.format(year_formula_first_param)

    # Finally, write to the YEAR column
    raw_data_sht.cell(row=row_to_write, column=RDS_YEAR_COL).value = year_formula
    raw_data_sht.cell(row=row_to_write, column=RDS_YEAR_COL).number_format = '0' # This line sets the formatting of the cell


    # END OF LOOP ITERATION FOR:
    #   for row in marketplace1_reader:
    # Finished writing data to this row. Now increment it by 1 so that the next iteration will write on the next row.
    row_to_write += 1
  # # # # # # # # # # END OF PROCESSING THE FIRST MARKET PLACE # # # # # # # # # #

  # TODO: PROCESS THE SECOND MARKETPLACE
  # # # # # # # # # # PROCESS THE SECOND MARKET PLACE # # # # # # # # # #



  # Save the file. IMPORTANT: You MUST save the file to see changes.
  master_wb.save(master_wb_path)
  
  # When we're all done, close master workbook
  master_wb.close


if __name__ == "__main__":
  main()
